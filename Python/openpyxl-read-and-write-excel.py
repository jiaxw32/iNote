r"""
OpenPyXL: https://openpyxl.readthedocs.io/en/stable/

install openpyxl:  `python -m pip install openpyxl`

"""

import os
from openpyxl import load_workbook

sheet_map = {"语数外考号": 0, "物理考号": 1, "化学考号": 2, "生物考号": 3, "历史考号": 4, "地理考号": 5, "政治考号": 6}

def read_exam_number(filename: str):
    # 打开文件
    workBook = load_workbook(filename, read_only=True)
    # print all sheet names
    print(workBook.sheetnames)

    students_map = dict()

    for sheet in workBook:
        # get sheet name
        sheet_name = sheet.title
        print(f"sheet name: {sheet_name}")
        
        # min_row、max_row、max_col: 1-based index
        for row in sheet.iter_rows(min_row=2, max_col=2, max_row=145, values_only=True):
            # the type of row is `tuple class`
            # print(type(row), row, len(row))
            
            student_name = row[0] # name
            exam_number = row[1]  # exam number

            if student_name is None or exam_number is None:
                continue

            course_idx = sheet_map[sheet_name]

            if student_name not in students_map:
                course_lst = ["-1"] * 7 # init list with '-1'
                course_lst[course_idx] = exam_number
                students_map[student_name] = course_lst
            else:
                students_map[student_name][course_idx] = exam_number
    
    workBook.close()
    
    print(f"students count: {len(students_map)}")
    for k, v in  students_map.items():
        print(k, v)
    return students_map

def write_exam_number(filename: str, exam_number: dict):
    # open workbook
    wb = load_workbook(filename)
    # get the active sheet
    sheet = wb.active

    for row in sheet.iter_rows(min_row=3, max_col=30, max_row=145):
        name = row[4].value # student name
        if name is None:
            continue
        
        numbers = exam_number[name] # 考号列表
        print(f"student name: {name}, exam numbers: {numbers}")

        start_col = 19 # 语文, T 列，ord('T')-ord('A')
        col_offset = 0 # column offset
        for idx in range(0, len(numbers)):
            code = numbers[idx]
            
            if code == "-1": # 未选课程，不需要更新
                col_offset += 1
                continue
            
            if idx == 0:
                row[start_col].value = code # 语文
                row[start_col + 1].value = code # 数学
                row[start_col + 2].value = code # 英语
                col_offset = 3
            else:
                row[start_col + col_offset].value = code # 物理、化学、生物、历史、地理、政治
                col_offset += 1

    wb.save("./2022-2023学年第一学期期中高三名单-考号.xlsx")

if __name__ == '__main__':    
    
    filename = "./期中考场安排修改1.xlsx"
    if os.path.exists(filename):
        # read exam numbers from the source file
        exam_number = read_exam_number(filename)
        
        # write exam numbers to the target file
        students_file = "./2022-2023学年第一学期期中高三名单.xlsx"
        if os.path.exists(students_file):
            write_exam_number(students_file, exam_number)